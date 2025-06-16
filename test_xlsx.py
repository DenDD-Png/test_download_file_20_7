from openpyxl import load_workbook

workbook = load_workbook("tmp/")

sheet = workbook.active

#Печать информации из файла ексель
print(sheet.cell(row=2, column=3 ).value)

